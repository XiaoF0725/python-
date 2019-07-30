from fileinput import filename

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import logging
import os

start = datetime.datetime.today()
wb = Workbook()
# print(wb.sheetnames)
#wb.remove_sheet(wb.get_sheet_by_name("sheet"))
ws1 = wb.create_sheet("Mysheet",0)
ws1.title='自动化1'
ws2 = wb.create_sheet("Mysheet",1)
ws2.title='自动化2'
ws1['A1'] = "账号"
ws1['B1'] = "密码"
ft = Font(bold=True)
ws1['A1'].font = ft
ws1['B1'].font = ft
col = ws1.column_dimensions['A']
col.width = 17
ws1.column_dimensions['B'].width = 17
i = 2
j = 8451252630001
k = 8451252630001
while i <= 1001:
  ws1.cell(i,1).value = str(j)
  ws1.cell(i,2).value = 'Faxuan.%1234'
  i = i+1
  j = j+1
  for n in range (1000):
      k += n
ws2['A1'].font = ft
ws2['A1'] = '平均值'
ws2['B1'] = str(k/1000)
#ws2['B1'] = "=AVERAGE(自动化1!A2:自动化1!A1001)"
wb.save(start.date().strftime('%Y%m%d')+'.xlsx')
logging.basicConfig(filename='test_log.log',level=logging.DEBUG)
# 如果存在filename 则先删掉且记录日志
filepath = excel_path + '/' + filename
      if os.path.exists(filepath):
           os.remove(filepath)
           logging.info("文件删除:{}".format(filepath))
end =datetime.datetime.today()
print('运行时间: %s Seconds'%(end-start))
